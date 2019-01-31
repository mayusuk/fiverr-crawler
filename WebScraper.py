from bs4 import BeautifulSoup
from requests import get
from requests.exceptions import RequestException
from contextlib import closing
import json
import base64
import time
import re
import pandas as pd
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook


class OrderedDefaultDict(OrderedDict):
    def __missing__(self, key):
        value = list()
        self[key] = value
        return value


def get_page(url):
    time.sleep(15)
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:64.0) Gecko/20100101 Firefox/64.0'
        }
        vid = 'vid'
        uid = 'uid'
        data = 'response'
        response = json.dumps({'r': data, 'v': vid, 'u': uid})
        response = base64.b64encode(response.encode('utf-8'))

        cookies = {'_pxCaptcha': 'eyJyIjoiMDNBRjZqRHFWTE05Z0NOVXB3dC1leElDU0dDcFpFWGpKczR1RkJja0pjcWpmN1l5Tk1SckpMS2s3TnlkLVlEanA5anpHYlIzZkNnRE1jSTB1cmtZeEpPYUV3OGxlSG9qY3JWcmtrbjZwVzFsLTg2SGJnMWtzUnVPbEhpaVFLN3J0bU1OOEg2ZG9XU2ctV0xIMVhkT0ZaRWlwWHR3UzVOT3ZJX2t6V2NhQ3kxUV9yeW00UjNKR1F6Q24tYXBnNWlvZ0otUXhpN3ZkU0tzd1N6V0Z4LXpQdTctb0NjMjlfN3dtcF8xbWdFRXJYOGtQd1gwVHNHcDBlWEg5WGY4MlZQYTRxWl93bGVUc0hJNDdOUlZLa3YwSnhwWkNGb0lkdmt4QWdxeTY0NEJmYVZOLUpYcS1Ra2xpYkNyRWRBVU51dENpRU01d25CdjR3Sjc2SiIsInYiOiIiLCJ1IjoiIn0=; expires=Tue, 29 Jan 2019 05:15:34 GMT; path=/; domain=.fiverr.com'}
        with closing(get(url, stream=True, headers=headers, cookies=cookies)) as resp:
            if is_good_response(resp):
                return resp.content
            else:
                return None

    except RequestException as e:
        log_error('Error during requests to {0} : {1}'.format(url, str(e)))
        return None


def is_good_response(resp):

    content_type = resp.headers['Content-Type'].lower()
    return (resp.status_code == 200
            and content_type is not None)


def log_error(e):
    print(e)


def _json_object_hook(d, freelancers):

    gigs = d.pop("gigs", None)
    for gig in gigs:
        gig.pop("image_data", None)
        gig.pop("assets", None)
        gig.pop("impression_data", None)
        gig.pop("gig_image", None)
        if gig["seller_id"] not in freelancers:
            freelancers[gig["seller_id"]] = gig["seller_name"]

    return gigs


def get_gigs_from_api(url, api, categoryId, subcategoryId, page, freelancers, gigs):

    apiUrl = "{0}{1}.json?" \
             "category_id={2}&context_referrer=subcategory_listing" \
             "&filter=rating&host=subcategory" \
             "&sub_category_id={3}&page={4}"\
        .format(url, api, categoryId, subcategoryId, page)
    print("Crawling {0}".format(apiUrl))
    response = get_page(apiUrl)
    if response:
        jsonresponse = json.loads(response)
        gigs.extend(_json_object_hook(jsonresponse, freelancers))
        if jsonresponse["pagination"]["current_page"] == jsonresponse["pagination"]["number_of_pages"]:
            return
        get_gigs_from_api(url, api, categoryId, subcategoryId, page + 1, freelancers, gigs)
    return


def get_all_reviews(url, freelancerId, as_buyer=True):
    review_type = "as_buyer"
    if not as_buyer:
        review_type = "as_seller"
    api = "{0}/ratings/index?user_id={1}&page_size={2}&{3}=true".format(url, freelancerId, 100000, review_type)
    print("Crawling {0}".format(api))
    reviews = get_page(api)
    if reviews:
        reviews = json.loads(reviews)
        if "reviews" in reviews:
            return reviews["reviews"]
    return None


def get_freelancers_details(url, freelancerName):
    api = "{0}/{1}?source=gig-cards".format(url, freelancerName)
    print("Crawling {0}".format(api))
    page = get_page(api)
    response = {}
    user = {}
    user_found = response_found = testdata_found = False
    if page:
        page = BeautifulSoup(page, 'html.parser')
        script = page.find_all("script")

        for tag in script:
            if "window.initialData.SellerCard" in tag.get_text():
                content = tag.get_text()
                expression = "window.initialData.SellerCard = \{(.*)\};"
                matches = re.search(expression, content)
                user = matches.group()
                user = user.lstrip()
                user = user.replace("window.initialData.SellerCard = ", "")
                user = user.rstrip(";")
                user = json.loads(user)
                user_found = True

            if "document.viewSellerProfile" in tag.get_text():
                content = tag.get_text()
                expression = "document.viewSellerProfile = \{(.*)\};"
                matches = re.search(expression, content)
                response = matches.group()
                response = response.lstrip()
                response = response.replace("document.viewSellerProfile = ", "")
                response = response.rstrip(";")
                response = json.loads(response)
                response_found = True

            if "document.sellerTestsData" in tag.get_text():
                content = tag.get_text()
                expression = "document.sellerTestsData = \{(.*)\}"
                matches = re.search(expression, content)
                testdata = matches.group()
                testdata = testdata.lstrip()
                testdata = testdata.replace("document.sellerTestsData = ", "")
                testdata = testdata.rstrip(";")
                testdata = json.loads(testdata)
                testdata_found = True

    if response_found:
        if user_found:
            response["user"] = user["user"]
        if testdata_found:
            response["testdata"] = testdata["test_results"]
        return response
    return {}


def write_to_excel(sheetname, dataframe, writer):

    df = pd.DataFrame(dataframe)

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=sheetname)

def append_to_excel(filename, sheet_name, df, startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    df = pd.DataFrame(df)
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()



if __name__ == '__main__':

    start = time.time()

    url = 'https://www.fiverr.com'
    excel_file = 'pandas_simple.xlsx'
    categoriesFile = open("categories", "w")
    categoriesFile.write("category, categoryId\n")
    subCategoriesFile = open("subcategories", "w")
    subCategoriesFile.write("categoryId, subcategory, subcategoryId\n")
    fiverrUrlFile = open("FiverrUrls", "r")
    fiverrUrls = json.load(fiverrUrlFile)
    gigs = defaultdict()
    freelancers = {}
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    # writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    #
    # df = pd.DataFrame(freelancers)
    #
    # # Convert the dataframe to an XlsxWriter Excel object.
    # df.to_excel(writer, sheet_name="test")
    # writer.save()
    # writer.close()


    subcategory_dataframe = OrderedDefaultDict()
    category_dataframe= OrderedDefaultDict()
    for menu in fiverrUrls["menu"]:
        if menu["type"] == "categories":
            for category in menu["categories"]:
                categoriesFile.write("{0},{1}\n".format(category["name"], category["id"]))
                category_dataframe["category"].append(category["name"])
                category_dataframe["category_id"].append(category["id"])
                if category["name"]  == "Programming   Tech":

                    for subcategory in category["subcategories"]:
                        gig_list = []
                        subCategoriesFile.write("{0},{1},{2}\n".format(category["id"], subcategory["name"], subcategory["id"]))
                        get_gigs_from_api(url, subcategory["url"], category["id"], subcategory["id"], 0, freelancers, gig_list)
                        gigs[subcategory["id"]] = gig_list
                        subcategory_dataframe["categoryId"].append(category["id"])
                        subcategory_dataframe["subcategory"].append(subcategory["name"])
                        subcategory_dataframe["subcategoryId"].append(subcategory["id"])
    fiverrUrlFile.close()
    append_to_excel(excel_file, "categories", category_dataframe)
    append_to_excel(excel_file, "subcategories", subcategory_dataframe)

    gigsFile = open("gigs", "w")
    gigsFile.write("subcategoryId|categoryId|gig_id|title|status|price|rating|rating_count|"
                   "is_featured|gig_created|gig_locale|max_qantity|seller_id|seller_country\n")
    gigs_data_frame = defaultdict(list)
    unique_gigs = defaultdict()

    gigs_package_dataframe = OrderedDefaultDict()

    for key, values in gigs.items():
        for value in values:
            gigsFile.write("{0}|{1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}|{9}|{10}|{11}|{12}|{13}\n".format(key, value["category_id"], value["gig_id"],
                                                                value["title"], value["status"], value["price"],
                                                                value["rating"], value["rating_count"], value["is_featured"],
                                                                value["gig_created"], value["gig_locale"], value["max_quantity"],
                                                                value["seller_id"], value["seller_country"]))
            if value["gig_id"] not in unique_gigs:
                skill_list = ""
                if value.get("skills", None) :
                    for skill in value.get("skills", None):
                        skill_list += "," + skill
                    skill_list = skill_list.lstrip(",")
                unique_gigs[value["gig_id"]] = True
                gigs_data_frame["subcategoryId"].append(key)
                gigs_data_frame["categoryId"].append(value.get("category_id", None))
                gigs_data_frame["gig_id"].append(value.get("gig_id", None))
                gigs_data_frame["title"].append(value.get("title", None))
                gigs_data_frame["status"].append(value.get("status", None))
                gigs_data_frame["price"].append(value.get("price",None))
                gigs_data_frame["rating"].append(value.get("rating", None))
                gigs_data_frame["rating_count"].append(value.get("rating_count", None))
                gigs_data_frame["is_featured"].append(value.get("is_featured", None))
                gigs_data_frame["gig_created"].append(value.get("gig_created", None))
                gigs_data_frame["gig_locale"].append(value.get("gig_locale", None))
                gigs_data_frame["max_qantity"].append(value.get("max_quantity", None))
                gigs_data_frame["skills"].append(skill_list)
                gigs_data_frame["seller_id"].append(value.get("seller_id", None))
                gigs_data_frame["seller_country"].append(value.get("seller_country", None))
                gigs_data_frame["is_new_seller"].append(value.get("is_new_seller", None))
                gigs_data_frame["seller_avg_response"].append(value.get("seller_avg_response", None))
                gigs_data_frame["seller_level"].append(value.get("seller_level", None))
                gigs_data_frame["price_highest"].append(value.get("price_highest", None))
                gigs_data_frame["gig_url"].append(value.get("gig_url", None))

                if value.get("packages", None):
                    packages = value.get("packages", None)
                    for package in packages:
                        gigs_package_dataframe["gig_id"].append(value.get("gig_id", None))
                        gigs_package_dataframe["seller_id"].append(value.get("seller_id", None))
                        gigs_package_dataframe["package_id"].append(package.get("package_id", None))
                        gigs_package_dataframe["updated_at"].append(package.get("updated_at", None))
                        gigs_package_dataframe["title"].append(package.get("title", None))
                        gigs_package_dataframe["description"].append(package.get("description", None))
                        gigs_package_dataframe["duration"].append(package.get("description", None))
                        gigs_package_dataframe["duration_unit"].append(package.get("duration_unit", None))
                        gigs_package_dataframe["price"].append(package.get("price", None))
                        gigs_package_dataframe["content"].append(package.get("content", None))


    # append_to_excel(excel_file, "gigs", gigs_data_frame)
    #
    # gigsFile.close()
    #
    # freelancerFile = open("freelancersList", "w")
    # freelancerFile.write("seller_id,seller_name\n")
    # for seller_id, seller_name in freelancers.items():
    #     freelancerFile.write("{0},{1}\n".format(seller_id, seller_name))
    # freelancerFile.close()
    #
    # reviews_as_buyer = defaultdict(list)
    # reviews_as_seller = defaultdict(list)
    #
    # freelancerFile = open("freelancersList", "r")
    # freelancerFile.readline()
    # for freelancer in freelancerFile.readlines():
    #     freelancerId = freelancer.split(",")[0]
    #     # get_freelancers_details(url, freelancerId)
    #     response = get_all_reviews(url, freelancerId, as_buyer=True)
    #     if response:
    #         reviews_as_buyer[freelancerId]= response
    #     response = get_all_reviews(url, freelancerId, as_buyer=False)
    #     if response:
    #         reviews_as_seller[freelancerId] = response
    #
    # reviews_as_buyer_dataframe = OrderedDefaultDict()
    # reviews_as_seller_dataframe = OrderedDefaultDict()
    #
    # reviews_as_buyer_file = open("BuyerReviews", "w")
    # reviews_as_seller_file = open("SellerReviews", "w")
    # reviews_as_buyer_file.write("freelancerId|reviewer_username|rating|comment|created_at\n")
    # reviews_as_seller_file.write("freelancerId|reviewer_username|rating|comment|created_at\n")
    # for freelancerId,reviews in reviews_as_buyer.items():
    #     for review in reviews:
    #         reviews_as_buyer_file.write("{0}|{1}|{2}|{3}|{4}\n".format(freelancerId, review["username"],
    #                                                                    review["value"],
    #                                                                    review["comment"],
    #                                                                    review["created_at"]))
    #         reviews_as_buyer_dataframe["freelancerId"].append(freelancerId)
    #         reviews_as_buyer_dataframe["reviewer_username"].append(review["username"])
    #         reviews_as_buyer_dataframe["rating"].append(review["value"])
    #         reviews_as_buyer_dataframe["comment"].append(review["comment"])
    #         reviews_as_buyer_dataframe["created_at"].append(review["created_at"])
    #
    # append_to_excel(excel_file, "reviews_as_buyer", reviews_as_buyer_dataframe)
    #
    # for freelancerId, reviews in reviews_as_seller.items():
    #     for review in reviews:
    #         reviews_as_seller_file.write("{0}|{1}|{2}|{3}|{4}\n".format(freelancerId, review["username"],
    #                                            review["value"],
    #                                            review["comment"],
    #                                            review["created_at"]))
    #         reviews_as_seller_dataframe["freelancerId"].append(freelancerId)
    #         reviews_as_seller_dataframe["reviewer_username"].append(review["username"])
    #         reviews_as_seller_dataframe["rating"].append(review["value"])
    #         reviews_as_seller_dataframe["comment"].append(review["comment"])
    #         reviews_as_seller_dataframe["created_at"].append(review["created_at"])
    #
    # append_to_excel(excel_file, "reviews_as_seller", reviews_as_seller_dataframe)
    #
    # reviews_as_seller_file.close()
    # reviews_as_buyer_file.close()
    #
    # freelancersDetails_dataframe = OrderedDefaultDict()
    # freelancerEdu_dataframe = OrderedDefaultDict()
    # freelancerCert_dataframe = OrderedDefaultDict()
    # freelancertests_dataframe = OrderedDefaultDict()
    #
    # freelancersDetails = defaultdict()
    # freelancerFile = open("freelancersList", "r")
    # freelancerFile.readline()
    # for line in freelancerFile.readlines():
    #     freelancerUserName = line.split(",")[1]
    #     freelancerUserName = freelancerUserName.rstrip("\n")
    #     freelancersDetails[freelancerUserName] = get_freelancers_details(url, freelancerUserName)
    # freelancersDetailsFile = open("freelancersDetails", "w")
    # freelancersDetailsFile.write("user_id|username|rating|ratings_count|"
    #                              "country|member_since|is_pro|is_seller|is_pro_experience|"
    #                              "is_ambassador|custom_orders_allowed|active_skills|languages\n")
    # for freelancerUserName, data in freelancersDetails.items():
    #     user_data = data.get("user", None)
    #     if user_data is None:
    #         print("{0} has empty data".format(freelancerUserName))
    #     if user_data:
    #
    #         skills = data.get("skills", None)
    #         active_skills = ""
    #         if skills:
    #             for skill in skills["list"]:
    #                 if skill["status"] == "active":
    #                     active_skills += "," + skill["name"]
    #         active_skills = active_skills.lstrip(",")
    #
    #         languges = data.get("proficient_languages", None)
    #         proficient_languages = ""
    #
    #         if languges:
    #             for languge in languges["list"]:
    #                 proficient_languages += "," + languge["name"]
    #         proficient_languages = proficient_languages.lstrip(",")
    #
    #         social_accounts = ""
    #
    #         if "social_accounts" in data:
    #             accounts = data.get("social_accounts", None)
    #             for account in accounts["list"]:
    #                 social_accounts += "," + account["value"]
    #         social_accounts.lstrip(",")
    #
    #         if data.get("testdata") is not None and data.get("testdata") :
    #             for testdata in data["testdata"]:
    #                 freelancertests_dataframe["userid"].append(user_data.get("id", None))
    #                 freelancertests_dataframe["test_title"].append(testdata.get("title", None))
    #                 freelancertests_dataframe["score"].append(testdata.get("score", None))
    #                 freelancertests_dataframe["platform_name"].append(testdata.get("platform_name", None))
    #                 freelancertests_dataframe["passed"].append(testdata.get("passed", None))
    #                 freelancertests_dataframe["total_questions"].append(testdata.get("total_questions", None))
    #                 freelancertests_dataframe["slug"].append(testdata.get("slug", None))
    #                 freelancertests_dataframe["status"].append(testdata.get("status", None))
    #
    #         if "certifications" in data:
    #             for cert in data["certifications"]["list"]:
    #                 freelancerCert_dataframe["userid"].append(user_data.get("id", None))
    #                 freelancerCert_dataframe["certification_name"].append(cert.get("certification_name",None))
    #                 freelancerCert_dataframe["received_from"].append(cert.get("received_from",None))
    #                 freelancerCert_dataframe["year"].append(cert.get("year",None))
    #
    #
    #         education = []
    #         if "educations" in data :
    #             education = data["educations"]["list"]
    #             for edu in education:
    #                 freelancerEdu_dataframe["userid"].append(user_data.get("id", None))
    #                 freelancerEdu_dataframe["degree"].append(edu.get("degree",None))
    #                 freelancerEdu_dataframe["from_year"].append(edu.get("from_year",None))
    #                 freelancerEdu_dataframe["to_year"].append(edu.get("to_year",None))
    #                 freelancerEdu_dataframe["degree_title"].append(edu.get("degree_title",None))
    #                 freelancerEdu_dataframe["school"].append(edu.get("school",None))
    #                 freelancerEdu_dataframe["country"].append(edu.get("country",None))
    #
    #
    #         freelancersDetailsFile.write("{0}|{1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}|{9}|{10}|{11}|{12}\n".format(user_data["id"],
    #                                                                             freelancerUserName,
    #                                                                             user_data["rating"],
    #                                                                             user_data["ratings_count"],
    #                                                                             user_data["country"],
    #                                                                             user_data["member_since"],
    #                                                                             user_data["is_pro"],
    #                                                                             user_data["is_seller"],
    #                                                                             user_data["is_pro_experience"],
    #                                                                             user_data["is_ambassador"],
    #                                                                             user_data["custom_orders_allowed"],
    #                                                                             active_skills,
    #                                                                             proficient_languages))
    #         freelancersDetails_dataframe["user_id"].append(user_data.get("id", None))
    #         freelancersDetails_dataframe["username"].append(freelancerUserName)
    #         freelancersDetails_dataframe["overview"].append(data.get('overview', None))
    #         freelancersDetails_dataframe["rating"].append(user_data.get("rating", None))
    #         freelancersDetails_dataframe["ratings_count"].append(user_data.get("ratings_count", None))
    #         freelancersDetails_dataframe["country"].append(user_data.get("country", None))
    #         freelancersDetails_dataframe["member_since"].append(user_data.get("member_since", None))
    #         freelancersDetails_dataframe["is_pro"].append(user_data.get("is_pro", None))
    #         freelancersDetails_dataframe["is_seller"].append(user_data.get("is_seller", None))
    #         freelancersDetails_dataframe["is_pro_experience"].append(user_data.get("is_pro_experience", None))
    #         freelancersDetails_dataframe["is_ambassador"].append(user_data.get("is_ambassador", None))
    #         freelancersDetails_dataframe["custom_orders_allowed"].append( user_data.get("custom_orders_allowed", None))
    #         freelancersDetails_dataframe["active_skills"].append(active_skills)
    #         freelancersDetails_dataframe["languages"].append(proficient_languages)
    #         freelancersDetails_dataframe["social_accounts"].append(social_accounts)
    #
    # arg = {'header': False}
    # append_to_excel(excel_file, "freelancers", freelancersDetails_dataframe, **arg)
    # append_to_excel(excel_file, "freelancers_education", freelancerEdu_dataframe, **arg)
    # append_to_excel(excel_file, "freelancers_cert", freelancerCert_dataframe, **arg)
    # append_to_excel(excel_file, "freelancers_tests", freelancertests_dataframe, **arg)


    print("Total time taken {0}".format(time.time() - start))
