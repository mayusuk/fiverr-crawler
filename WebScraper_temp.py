from bs4 import BeautifulSoup
from requests import get
from requests.exceptions import RequestException
from contextlib import closing
import json
import base64
import time
import re
import pandas as pd
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook
import argparse
import demjson

class OrderedDefaultDict(OrderedDict):
    def __missing__(self, key):
        value = list()
        self[key] = value
        return value


def get_page(url):
    time.sleep(3)
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:64.0) Gecko/20100101 Firefox/64.0'
        }
        vid = 'vid'
        uid = 'uid'
        data = 'response'
        response = json.dumps({'r': data, 'v': vid, 'u': uid})
        response = base64.b64encode(response.encode('utf-8'))

        cookies = {'_pxCaptcha': 'eyJyIjoiMDNBRjZqRHFXWnpXTWVldVdQUXlWcGZhYVZPaTBna1NqSTl1UWN4U3BwYmM2NU9HVVhPSWRPSkV5UUUxT2IwZlY5ZlU5MHo2VVNmeTl4RzUwMkQxSHQwVW5ib1FFeVdoSFRITDM1VUZ6clRtMWk3dVpucENjbFc5ZGFmNldmd1EtUjVNUjdqdzJocDUtYXFObktyYUdTLVVUbktTRWtva2ZlOGp6V2dtMWw5X04zS1RUUFZLbDJUN1RDSkR1NUZreEt0T2RReU9rMk9rbHJJOGxqVG00dzhfejBZdzducUpoekVYTUtTQ01OOUFja092ZXVBM2dMUnk2djE3VldVNEJuNUszcXYwUUNQU195Tkx4aGZSOE9JYkU3V3BtQUE2a3U5Z21TS1czbTFidWdUMk5GRUhhaGtWRy14OHVwU3VxWW9UaHVwclJ0QVBTUSIsInYiOiIiLCJ1IjoiIn0=; expires=Tue, 29 Jan 2019 05:15:34 GMT; path=/; domain=.fiverr.com'}
        with closing(get(url, stream=True, headers=headers, cookies=cookies)) as resp:
            if is_good_response(resp):
                return resp.content
            else:
                return None

    except RequestException as e:
        log_error('Error during requests to {0} : {1}'.format(url, str(e)))
        return None


def is_good_response(resp):

    content_type = resp.headers['Content-Type'].lower()
    return (resp.status_code == 200
            and content_type is not None)


def log_error(e):
    print(e)


def _json_object_hook(d, freelancers):

    gigs = d.pop("gigs", None)
    for gig in gigs:
        gig.pop("image_data", None)
        gig.pop("assets", None)
        gig.pop("impression_data", None)
        gig.pop("gig_image", None)
        if gig["seller_id"] not in freelancers:
            freelancers[gig["seller_id"]] = gig["seller_name"]

    return gigs


def get_gigs_from_api(url, api, categoryId, subcategoryId, page, freelancers, gigs):

    apiUrl = "{0}{1}.json?" \
             "category_id={2}&context_referrer=subcategory_listing" \
             "&filter=rating&host=subcategory" \
             "&sub_category_id={3}&page={4}"\
        .format(url, api, categoryId, subcategoryId, page)
    print("Crawling {0}".format(apiUrl))
    response = get_page(apiUrl)
    if response:
        jsonresponse = json.loads(response)
        gigs.extend(_json_object_hook(jsonresponse, freelancers))
        if jsonresponse["pagination"]["current_page"] == jsonresponse["pagination"]["number_of_pages"]:
            return
        get_gigs_from_api(url, api, categoryId, subcategoryId, page + 1, freelancers, gigs)
    return

def get_gig_details(url, suburl):
    apiUrl = "{0}{1}".format(url, suburl)
    print("Crawling {0}".format(apiUrl))
    response = get_page(apiUrl)
    if response:

        page = BeautifulSoup(response, 'html.parser')
        script = page.find_all("script")
        gigdata = None
        for tag in script:
            if "gigData = {" in tag.get_text():
                content = tag.get_text()
                content = content.replace("\n", "")
                content = content.replace("\r", "")
                expression = "gigData = \{(.*)\},"
                matches = re.search(expression, content)
                gigdata = matches.group()
                gigdata = gigdata.lstrip()
                gigdata = gigdata.replace("gigData = ", "")
                gigdata = gigdata.rstrip(",")
                gigdata = demjson.decode(gigdata)
                break

        main_desc = page.find("div", {"class": "gig-main-desc"})
        if main_desc:
            main_desc = main_desc.get_text()
        return gigdata, main_desc



def get_all_reviews(url, freelancerId, positive=True):
    review_type = "positive"
    if not positive:
        review_type = "negative"
    api = "{0}/ratings/index?gig_id={1}&page_size={2}&type={3}".format(url, freelancerId, 100000, review_type)
    print("Crawling {0}".format(api))
    reviews = get_page(api)
    if reviews:
        reviews = json.loads(reviews)
        if "reviews" in reviews:
            return reviews["reviews"]
    return None


def get_freelancers_details(url, freelancerName):
    api = "{0}/{1}?source=gig-cards".format(url, freelancerName)
    print("Crawling {0}".format(api))
    page = get_page(api)
    response = {}
    user = {}
    user_found = response_found = testdata_found = False
    if page:
        page = BeautifulSoup(page, 'html.parser')
        script = page.find_all("script")

        for tag in script:
            if "window.initialData.SellerCard" in tag.get_text():
                content = tag.get_text()
                expression = "window.initialData.SellerCard = \{(.*)\};"
                matches = re.search(expression, content)
                user = matches.group()
                user = user.lstrip()
                user = user.replace("window.initialData.SellerCard = ", "")
                user = user.rstrip(";")
                user = json.loads(user)
                user_found = True

            if "document.viewSellerProfile" in tag.get_text():
                content = tag.get_text()
                expression = "document.viewSellerProfile = \{(.*)\};"
                matches = re.search(expression, content)
                response = matches.group()
                response = response.lstrip()
                response = response.replace("document.viewSellerProfile = ", "")
                response = response.rstrip(";")
                response = json.loads(response)
                response_found = True

            if "document.sellerTestsData" in tag.get_text():
                content = tag.get_text()
                expression = "document.sellerTestsData = \{(.*)\}"
                matches = re.search(expression, content)
                testdata = matches.group()
                testdata = testdata.lstrip()
                testdata = testdata.replace("document.sellerTestsData = ", "")
                testdata = testdata.rstrip(";")
                testdata = json.loads(testdata)
                testdata_found = True

    if response_found:
        if user_found:
            response["user"] = user["user"]
        if testdata_found:
            response["testdata"] = testdata["test_results"]
        return response
    return {}


def write_to_excel(sheetname, dataframe, writer):

    df = pd.DataFrame(dataframe)

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=sheetname)


def append_to_excel(filename, sheet_name, df, startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    df = pd.DataFrame(df)
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def crawl_gigs_by_category(url, categoryName, excel_file):

    categoriesFile = open("categories", "w")
    categoriesFile.write("category, categoryId\n")
    subCategoriesFile = open("subcategories", "w")
    subCategoriesFile.write("categoryId, subcategory, subcategoryId\n")
    fiverrUrlFile = open("FiverrUrls", "r")
    fiverrUrls = json.load(fiverrUrlFile)
    gigs = defaultdict()
    freelancers = {}

    subcategory_dataframe = OrderedDefaultDict()
    category_dataframe= OrderedDefaultDict()
    for menu in fiverrUrls["menu"]:
        if menu["type"] == "categories":
            for category in menu["categories"]:
                categoriesFile.write("{0},{1}\n".format(category["name"], category["id"]))
                category_dataframe["category"].append(category["name"])
                category_dataframe["category_id"].append(category["id"])
                if category["name"] == categoryName:

                    for subcategory in category["subcategories"]:
                        gig_list = []
                        subCategoriesFile.write("{0},{1},{2}\n".format(category["id"], subcategory["name"], subcategory["id"]))
                        get_gigs_from_api(url, subcategory["url"], category["id"], subcategory["id"], 0, freelancers, gig_list)
                        gigs[subcategory["id"]] = gig_list
                        subcategory_dataframe["categoryId"].append(category["id"])
                        subcategory_dataframe["subcategory"].append(subcategory["name"])
                        subcategory_dataframe["subcategoryId"].append(subcategory["id"])
    fiverrUrlFile.close()
    append_to_excel(excel_file, "categories", category_dataframe)
    append_to_excel(excel_file, "subcategories", subcategory_dataframe)

    gigsFile = open("gigs", "w")
    gigsFile.write("subcategoryId|categoryId|gig_id|title|status|price|rating|rating_count|"
                   "is_featured|gig_created|gig_locale|max_qantity|seller_id|seller_country\n")
    gigs_data_frame = defaultdict(list)
    gigs_package_frame = OrderedDefaultDict()

    unique_gigs = defaultdict()
    for key, values in gigs.items():
        for value in values:
            gigsFile.write("{0}|{1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}|{9}|{10}|{11}|{12}|{13}\n".format(key, value["category_id"], value["gig_id"],
                                                                value["title"], value["status"], value["price"],
                                                                value["rating"], value["rating_count"], value["is_featured"],
                                                                value["gig_created"], value["gig_locale"], value["max_quantity"],
                                                                value["seller_id"], value["seller_country"]))
            gig_url = discription = None
            gigdata = {}
            if value["gig_id"] not in unique_gigs:
                skill_list = ""
                if value.get("skills", None) :
                    for skill in value.get("skills", None):
                        skill_list += "," + skill
                    skill_list = skill_list.lstrip(",")
                gig_url = value.get("gig_url", None)
                unique_gigs[value["gig_id"]] = True
                gigs_data_frame["subcategoryId"].append(key)
                gigs_data_frame["categoryId"].append(value.get("category_id", None))
                gigs_data_frame["gig_id"].append(value.get("gig_id", None))
                gigs_data_frame["title"].append(value.get("title", None))
                gigs_data_frame["status"].append(value.get("status", None))
                gigs_data_frame["price"].append(value.get("price",None))
                gigs_data_frame["rating"].append(value.get("rating", None))
                gigs_data_frame["rating_count"].append(value.get("rating_count", None))
                gigs_data_frame["is_featured"].append(value.get("is_featured", None))
                gigs_data_frame["fastest_delivery_time"].append(value.get("fastest_delivery_time", None))
                gigs_data_frame["avg_delivery_time"].append(value.get("avg_delivery_time", None))
                gigs_data_frame["gig_created"].append(value.get("gig_created", None))
                gigs_data_frame["gig_locale"].append(value.get("gig_locale", None))
                gigs_data_frame["max_qantity"].append(value.get("max_quantity", None))
                gigs_data_frame["skills"].append(skill_list)
                gigs_data_frame["seller_id"].append(value.get("seller_id", None))
                gigs_data_frame["seller_country"].append(value.get("seller_country", None))
                gigs_data_frame["is_new_seller"].append(value.get("is_new_seller", None))
                gigs_data_frame["seller_avg_response"].append(value.get("seller_avg_response", None))
                gigs_data_frame["seller_level"].append(value.get("seller_level", None))
                gigs_data_frame["price_highest"].append(value.get("price_highest", None))
                gigs_data_frame["gig_url"].append(value.get("gig_url", None))



                if gig_url:
                    gigdata, discription = get_gig_details(url, gig_url)
                gigs_data_frame["ordersInQueue"].append(gigdata.get("ordersInQueue", None))
                gigs_data_frame["tags"].append(gigdata.get("tags", None))
                gigs_data_frame["pricingModel"].append(gigdata.get("pricingModel", None))
                if gigdata.get("pricingModel", None) == 'Package' and value.get("packages", None):
                    for package in value.get("packages", []):
                        gigs_package_frame["gig_id"].append(value.get("gig_id", None))
                        gigs_package_frame["title"].append(package.get("title", None))
                        gigs_package_frame["description"].append(package.get("description", None))
                        gigs_package_frame["duration"].append(package.get("duration", None))
                        gigs_package_frame["duration_unit"].append(package.get("duration_unit", None))
                        gigs_package_frame["price"].append(package.get("price", None))
                        modifications = None
                        extra_fast_price = None
                        extra_fast_duration = None
                        if package.get("content", None):
                            for content in package.get("content", []):
                                if content.get('buyable_type', None) == "modifications" and content.get("extra_data", None):
                                    modifications = content.get("extra_data", None).get("included_modifications", None)
                                if content.get('buyable_type', None) == "extra_fast":
                                    extra_fast_price = content.get("price", None)
                                    extra_fast_duration = content.get("duration", None)
                        gigs_package_frame["modifications"].append(modifications)
                        gigs_package_frame["extra_fast_price"].append(extra_fast_price)
                        gigs_package_frame["extra_fast_duration"].append(extra_fast_duration)


    append_to_excel(excel_file, "gigs", gigs_data_frame)
    append_to_excel(excel_file, "gigs_package", gigs_package_frame)

    gigsFile.close()

    freelancerFile = open("freelancersList", "w")
    freelancerFile.write("seller_id,seller_name\n")
    for seller_id, seller_name in freelancers.items():
        freelancerFile.write("{0},{1}\n".format(seller_id, seller_name))
    freelancerFile.close()

def crawl_reviews(url, excel_file, start=0, end=200):

    i = 1
    start = int(start)
    end = int(end)
    positive_reviews = defaultdict(list)
    negative_reviews = defaultdict(list)

    header = False
    startrow = None
    if start == 1:
        header = True
        startrow = 0
    arg = {'header': header}

    gigs = open("gigs", "r")
    gigs.readline()
    for gig in gigs.readlines():

        if i >= start and i <= end:
            gig_id = gig.split("|")[2]
            gig_id = gig_id.lstrip("'")
            response = get_all_reviews(url, gig_id, positive=True)
            if response:
                positive_reviews[gig_id]= response
            response = get_all_reviews(url, gig_id, positive=False)
            if response:
                negative_reviews[gig_id] = response
        i += 1

    positive_reviews_dataframe = OrderedDefaultDict()
    negative_reviews_dataframe = OrderedDefaultDict()

    reviews_as_buyer_file = open("BuyerReviews", "w")
    reviews_as_seller_file = open("SellerReviews", "w")
    reviews_as_buyer_file.write("gig_id|reviewer_username|rating|comment|created_at\n")
    reviews_as_seller_file.write("gig_id|reviewer_username|rating|comment|created_at\n")
    for gig_id,reviews in positive_reviews.items():
        for review in reviews:
            reviews_as_buyer_file.write("{0}|{1}|{2}|{3}|{4}\n".format(gig_id, review["username"],
                                                                       review["value"],
                                                                       review["comment"],
                                                                       review["created_at"]))
            positive_reviews_dataframe["gig_id"].append(gig_id)
            positive_reviews_dataframe["reviewer_username"].append(review["username"])
            positive_reviews_dataframe["rating"].append(review["value"])
            positive_reviews_dataframe["comment"].append(review["comment"])
            positive_reviews_dataframe["created_at"].append(review["created_at"])
            positive_reviews_dataframe["work_sample"].append(review.get("work_sample", None))
            seller_response = review.get("seller_response", None)
            if seller_response:
                seller_response = seller_response.get("comment", None)
            else:
                seller_response = None
            positive_reviews_dataframe["seller_response"].append(seller_response)

    append_to_excel(excel_file, "positive_reviews", positive_reviews_dataframe, startrow=startrow, **arg)

    for gig_id, reviews in negative_reviews.items():
        for review in reviews:
            reviews_as_seller_file.write("{0}|{1}|{2}|{3}|{4}\n".format(gig_id, review["username"],
                                               review["value"],
                                               review["comment"],
                                               review["created_at"]))
            negative_reviews_dataframe["gig_id"].append(gig_id)
            negative_reviews_dataframe["reviewer_username"].append(review["username"])
            negative_reviews_dataframe["rating"].append(review["value"])
            negative_reviews_dataframe["comment"].append(review["comment"])
            negative_reviews_dataframe["created_at"].append(review["created_at"])
            negative_reviews_dataframe["work_sample"].append(review.get("work_sample", None))
            seller_response = review.get("seller_response", None)
            if seller_response:
                seller_response = seller_response.get("comment", None)
            else:
                seller_response = None
            negative_reviews_dataframe["seller_response"].append(seller_response)


    append_to_excel(excel_file, "negative reviews", negative_reviews_dataframe, startrow=startrow, **arg)

    reviews_as_seller_file.close()
    reviews_as_buyer_file.close()


def crawl_freelancers_details(url, excel_file, start=0, end=200):

    i = 1
    start = int(start)
    end = int(end)
    freelancersDetails_dataframe = OrderedDefaultDict()
    freelancerEdu_dataframe = OrderedDefaultDict()
    freelancerCert_dataframe = OrderedDefaultDict()
    freelancertests_dataframe = OrderedDefaultDict()

    freelancersDetails = defaultdict()
    freelancerFile = open("freelancersList", "r")
    freelancerFile.readline()
    for line in freelancerFile.readlines():
        if i >= start and i <= end:
            freelancerUserName = line.split(",")[1]
            freelancerUserName = freelancerUserName.rstrip("\n")
            freelancersDetails[freelancerUserName] = get_freelancers_details(url, freelancerUserName)
        i += 1
    freelancersDetailsFile = open("freelancersDetails", "w")
    freelancersDetailsFile.write("user_id|username|rating|ratings_count|"
                                 "country|member_since|is_pro|is_seller|is_pro_experience|"
                                 "is_ambassador|custom_orders_allowed|active_skills|languages\n")
    for freelancerUserName, data in freelancersDetails.items():
        user_data = data.get("user", None)
        if user_data is None:
            print("{0} has empty data".format(freelancerUserName))
        if user_data:

            skills = data.get("skills", None)
            active_skills = ""
            if skills:
                for skill in skills["list"]:
                    if skill["status"] == "active":
                        active_skills += "," + skill["name"]
            active_skills = active_skills.lstrip(",")

            languges = data.get("proficient_languages", None)
            proficient_languages = ""

            if languges:
                for languge in languges["list"]:
                    proficient_languages += "," + languge["name"]
            proficient_languages = proficient_languages.lstrip(",")

            social_accounts = ""

            if "social_accounts" in data:
                accounts = data.get("social_accounts", None)
                for account in accounts["list"]:
                    social_accounts += "," + account["value"]
            social_accounts.lstrip(",")

            if data.get("testdata") is not None and data.get("testdata"):
                for testdata in data["testdata"]:
                    freelancertests_dataframe["userid"].append(user_data.get("id", None))
                    freelancertests_dataframe["test_title"].append(testdata.get("title", None))
                    freelancertests_dataframe["score"].append(testdata.get("score", None))
                    freelancertests_dataframe["platform_name"].append(testdata.get("platform_name", None))
                    freelancertests_dataframe["passed"].append(testdata.get("passed", None))
                    freelancertests_dataframe["total_questions"].append(testdata.get("total_questions", None))
                    freelancertests_dataframe["slug"].append(testdata.get("slug", None))
                    freelancertests_dataframe["status"].append(testdata.get("status", None))

            if "certifications" in data:
                for cert in data["certifications"]["list"]:
                    freelancerCert_dataframe["userid"].append(user_data.get("id", None))
                    freelancerCert_dataframe["certification_name"].append(cert.get("certification_name", None))
                    freelancerCert_dataframe["received_from"].append(cert.get("received_from", None))
                    freelancerCert_dataframe["year"].append(cert.get("year", None))


            if "educations" in data:
                education = data["educations"]["list"]
                for edu in education:
                    freelancerEdu_dataframe["userid"].append(user_data.get("id", None))
                    freelancerEdu_dataframe["degree"].append(edu.get("degree", None))
                    freelancerEdu_dataframe["from_year"].append(edu.get("from_year", None))
                    freelancerEdu_dataframe["to_year"].append(edu.get("to_year", None))
                    freelancerEdu_dataframe["degree_title"].append(edu.get("degree_title", None))
                    freelancerEdu_dataframe["school"].append(edu.get("school", None))
                    freelancerEdu_dataframe["country"].append(edu.get("country", None))

            freelancersDetailsFile.write(
                "{0}|{1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}|{9}|{10}|{11}|{12}\n".format(user_data["id"],
                                                                                  freelancerUserName,
                                                                                  user_data["rating"],
                                                                                  user_data["ratings_count"],
                                                                                  user_data["country"],
                                                                                  user_data["member_since"],
                                                                                  user_data["is_pro"],
                                                                                  user_data["is_seller"],
                                                                                  user_data["is_pro_experience"],
                                                                                  user_data["is_ambassador"],
                                                                                  user_data["custom_orders_allowed"],
                                                                                  active_skills,
                                                                                  proficient_languages))
            freelancersDetails_dataframe["user_id"].append(user_data.get("id", None))
            freelancersDetails_dataframe["username"].append(freelancerUserName)
            freelancersDetails_dataframe["overview"].append(data.get('overview', None))
            freelancersDetails_dataframe["rating"].append(user_data.get("rating", None))
            freelancersDetails_dataframe["ratings_count"].append(user_data.get("ratings_count", None))
            freelancersDetails_dataframe["country"].append(user_data.get("country", None))
            freelancersDetails_dataframe["member_since"].append(user_data.get("member_since", None))
            freelancersDetails_dataframe["is_pro"].append(user_data.get("is_pro", None))
            freelancersDetails_dataframe["is_seller"].append(user_data.get("is_seller", None))
            freelancersDetails_dataframe["is_pro_experience"].append(user_data.get("is_pro_experience", None))
            freelancersDetails_dataframe["is_ambassador"].append(user_data.get("is_ambassador", None))
            freelancersDetails_dataframe["custom_orders_allowed"].append(user_data.get("custom_orders_allowed", None))
            freelancersDetails_dataframe["active_skills"].append(active_skills)
            freelancersDetails_dataframe["languages"].append(proficient_languages)
            freelancersDetails_dataframe["social_accounts"].append(social_accounts)

    header = False
    startrow= None
    if start == 1:
        header = True
        startrow = 0
    arg = {'header': header}

    append_to_excel(excel_file, "freelancers", freelancersDetails_dataframe, startrow=startrow, **arg)
    append_to_excel(excel_file, "freelancers_education", freelancerEdu_dataframe, startrow=startrow, **arg)
    append_to_excel(excel_file, "freelancers_cert", freelancerCert_dataframe, startrow=startrow, **arg)
    append_to_excel(excel_file, "freelancers_tests", freelancertests_dataframe, startrow=startrow, **arg)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument("--start", dest="start", help="Starting line number to start  processing", default=0)
    parser.add_argument("--end", dest="end", help="Last line number for  processing", default=200)
    parser.add_argument("--type", dest="type", help="type of the data to crawl", choices=["reviews", "freelancers"], required=True)
    args = parser.parse_args()

    start = time.time()

    url = 'https://www.fiverr.com'
    excel_file = 'pandas_simple.xlsx'

    # crawl_gigs_by_category(url, "Programming   Tech", excel_file)

    if args.type == "reviews":
        crawl_reviews(url, excel_file, args.start, args.end)

    if args.type == "freelancers":
        crawl_freelancers_details(url, excel_file, args.start, args.end)

    print("Total time taken {0}".format(time.time() - start))
